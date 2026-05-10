import json

import pandas as pd
import numpy as np
import pytest


def _build_synthetic_txn(n: int = 140) -> pd.DataFrame:
    dates = pd.date_range("2024-01-01", periods=n, freq="D")
    qty = [20 + (i % 7) + (i * 0.03) for i in range(n)]
    price = [100.0 + (i % 5) for i in range(n)]
    discount = [0.05 if i % 9 == 0 else 0.0 for i in range(n)]
    return pd.DataFrame(
        {
            "date": dates,
            "product_id": ["A"] * n,
            "category": ["cat1"] * n,
            "quantity": qty,
            "revenue": [qty[i] * price[i] * (1.0 - discount[i]) for i in range(n)],
            "price": price,
            "discount": discount,
            "promotion": [1.0 if i % 14 == 0 else 0.0 for i in range(n)],
            "freight_value": [5.0 + (i % 3) for i in range(n)],
            "cost": [65.0] * n,
            "stock": [999.0] * n,
            "weather_index": [i % 4 for i in range(n)],
        }
    )


def test_catboost_full_factor_mode_registered():
    import app

    assert "catboost_full_factors" in app.SCENARIO_CALC_MODES
    assert app.DEFAULT_SCENARIO_CALC_MODE == "enhanced_local_factors"


def test_dispatcher_accepts_catboost_full_factor_mode(monkeypatch):
    import app

    called = {"value": False, "factor_overrides": None}

    def fake_predict(**kwargs):
        called["value"] = True
        called["factor_overrides"] = kwargs.get("factor_overrides")
        return {
            "daily": pd.DataFrame({"actual_sales": [1.0], "revenue": [10.0], "profit": [2.0]}),
            "demand_total": 1.0,
            "revenue_total": 10.0,
            "profit_total": 2.0,
            "scenario_calc_mode": "catboost_full_factors",
        }

    monkeypatch.setattr(app, "predict_catboost_full_factor_projection", fake_predict)

    result = app.run_what_if_projection(
        trained_bundle={"daily_base": pd.DataFrame(), "future_dates": pd.DataFrame()},
        manual_price=100.0,
        factor_overrides={"factor__competitor_price": 120.0},
        scenario_calc_mode="catboost_full_factors",
    )

    assert called["value"] is True
    assert called["factor_overrides"] == {"factor__competitor_price": 120.0}
    assert result["scenario_calc_mode"] == "catboost_full_factors"


def test_legacy_and_enhanced_modes_still_resolve():
    import app

    assert app.resolve_scenario_calc_mode(None) == "enhanced_local_factors"
    assert app.resolve_scenario_calc_mode("legacy_current") == "legacy_current"
    assert app.resolve_scenario_calc_mode("enhanced_local_factors") == "enhanced_local_factors"


def test_unknown_scenario_calc_mode_raises():
    import app
    import pytest

    with pytest.raises(ValueError):
        app.run_what_if_projection(
            trained_bundle={"daily_base": pd.DataFrame(), "future_dates": pd.DataFrame()},
            manual_price=100.0,
            scenario_calc_mode="unknown_mode",
        )


def test_extra_factors_are_opt_in_only():
    from data_adapter import build_daily_from_transactions

    df = pd.DataFrame(
        {
            "date": pd.date_range("2024-01-01", periods=5),
            "product_id": ["A"] * 5,
            "quantity": [10, 12, 11, 13, 12],
            "revenue": [1000, 1200, 1100, 1300, 1200],
            "price": [100] * 5,
            "weather_index": [1, 2, 3, 4, 5],
        }
    )

    daily_default = build_daily_from_transactions(df, "A")
    assert "factor__weather_index" not in daily_default.columns

    daily_full = build_daily_from_transactions(df, "A", include_extra_factors=True)
    assert "factor__weather_index" in daily_full.columns


def test_legacy_and_enhanced_daily_base_has_no_extra_factor_columns():
    import app

    txn = _build_synthetic_txn()
    legacy = app.run_full_pricing_analysis_universal(
        normalized_txn=txn,
        target_category="cat1",
        target_sku="A",
        scenario_calc_mode="legacy_current",
    )
    enhanced = app.run_full_pricing_analysis_universal(
        normalized_txn=txn,
        target_category="cat1",
        target_sku="A",
        scenario_calc_mode="enhanced_local_factors",
    )

    legacy_cols = set((legacy["_trained_bundle"]["daily_base"]).columns)
    enhanced_cols = set((enhanced["_trained_bundle"]["daily_base"]).columns)
    assert not any(str(c).startswith("factor__") for c in legacy_cols)
    assert not any(str(c).startswith("factor__") for c in enhanced_cols)


def test_extra_factor_leakage_columns_are_blocked():
    from data_adapter import build_daily_from_transactions

    n = 20
    df = pd.DataFrame(
        {
            "date": pd.date_range("2024-01-01", periods=n),
            "product_id": ["A"] * n,
            "quantity": [10 + (i % 3) for i in range(n)],
            "revenue": [1000 + i * 10 for i in range(n)],
            "price": [100.0] * n,
            "weather_index": [i % 5 for i in range(n)],
            "competitor_price": [95 + (i % 4) for i in range(n)],
            "order_count": [100 + i for i in range(n)],
            "units_sold": [40 + i for i in range(n)],
            "actual_sales": [50 + i for i in range(n)],
        }
    )
    daily = build_daily_from_transactions(df, "A", include_extra_factors=True)
    cols = set(daily.columns)
    assert "factor__weather_index" in cols
    assert "factor__competitor_price" in cols
    assert "factor__order_count" not in cols
    assert "factor__units_sold" not in cols
    assert "factor__actual_sales" not in cols


def test_catboost_full_bundle_has_feature_report():
    from catboost_full_factor_engine import train_catboost_full_factor_bundle

    n = 90
    daily = pd.DataFrame(
        {
            "date": pd.date_range("2024-01-01", periods=n),
            "sales": [20 + (i % 7) + i * 0.1 for i in range(n)],
            "revenue": [2000 + i * 10 for i in range(n)],
            "price": [100 + (i % 5) for i in range(n)],
            "discount": [0.05 if i % 10 == 0 else 0.0 for i in range(n)],
            "promotion": [1.0 if i % 14 == 0 else 0.0 for i in range(n)],
            "freight_value": [5.0] * n,
            "cost": [60.0] * n,
            "stock": [999.0] * n,
            "factor__weather_index": [i % 4 for i in range(n)],
        }
    )
    future = pd.DataFrame({"date": pd.date_range("2024-04-01", periods=30)})

    bundle = train_catboost_full_factor_bundle(daily, future, min_train_days=60)

    assert "feature_report" in bundle
    assert "feature_cols" in bundle


def test_legacy_and_enhanced_modes_do_not_train_catboost(monkeypatch):
    import app

    def fail_train(*args, **kwargs):
        raise AssertionError("must not be called")

    monkeypatch.setattr(app, "train_catboost_full_factor_bundle", fail_train)
    txn = _build_synthetic_txn()

    result = app.run_full_pricing_analysis_universal(
        normalized_txn=txn,
        target_category="cat1",
        target_sku="A",
        scenario_calc_mode="enhanced_local_factors",
    )
    assert result.get("blocking_error", False) is False


def test_catboost_unavailable_returns_enhanced_fallback(monkeypatch):
    import app

    def fake_train(*args, **kwargs):
        return {"enabled": False, "reason": "catboost_unavailable", "warnings": ["CatBoost недоступен"], "feature_report": pd.DataFrame()}

    monkeypatch.setattr(app, "train_catboost_full_factor_bundle", fake_train)
    txn = _build_synthetic_txn()

    result = app.run_full_pricing_analysis_universal(
        normalized_txn=txn,
        target_category="cat1",
        target_sku="A",
        scenario_calc_mode="catboost_full_factors",
    )
    assert result.get("blocking_error", False) is False
    assert result.get("analysis_scenario_calc_mode") == "enhanced_local_factors"
    assert result.get("catboost_full_factor_attempt", {}).get("enabled") is False
    assert result.get("recommended_mode_status", {}).get("recommended_mode") == "enhanced_local_factors"


def test_catboost_full_mode_trains_only_when_selected(monkeypatch):
    import app

    calls = {"count": 0}
    orig = app.train_catboost_full_factor_bundle

    def wrapped(*args, **kwargs):
        calls["count"] += 1
        return orig(*args, **kwargs)

    monkeypatch.setattr(app, "train_catboost_full_factor_bundle", wrapped)
    txn = _build_synthetic_txn()
    app.run_full_pricing_analysis_universal(
        normalized_txn=txn,
        target_category="cat1",
        target_sku="A",
        scenario_calc_mode="catboost_full_factors",
    )
    assert calls["count"] == 1


def test_dispatcher_uses_bundle_mode_when_mode_not_passed(monkeypatch):
    import app

    called = {"value": False}

    def fake_predict(**kwargs):
        called["value"] = True
        return {"daily": pd.DataFrame(), "scenario_calc_mode": "catboost_full_factors"}

    monkeypatch.setattr(app, "predict_catboost_full_factor_projection", fake_predict)
    app.run_what_if_projection(
        trained_bundle={"analysis_scenario_calc_mode": "catboost_full_factors", "daily_base": pd.DataFrame(), "future_dates": pd.DataFrame()},
        manual_price=100.0,
    )
    assert called["value"] is True


def test_catboost_mode_summary_label_not_legacy():
    import app

    txn = _build_synthetic_txn(170)
    res = app.run_full_pricing_analysis_universal(
        normalized_txn=txn,
        target_category="cat1",
        target_sku="A",
        scenario_calc_mode="catboost_full_factors",
    )
    assert res.get("blocking_error", False) is False
    bundle = res["_trained_bundle"]
    wr = app.run_what_if_projection(
        bundle,
        manual_price=float(bundle["base_ctx"]["price"]) * 1.01,
        scenario_calc_mode="catboost_full_factors",
    )
    payload_blob, _ = app.build_manual_scenario_artifacts(res, wr)
    payload = json.loads(payload_blob.decode("utf-8"))
    assert str(payload.get("scenario_calc_mode")) == "catboost_full_factors"
    assert str(payload.get("legacy_or_enhanced_label")) == "catboost_full_factors"


def test_custom_factor_override_changes_catboost_prediction():
    from catboost_full_factor_engine import (
        predict_catboost_full_factor_projection,
        train_catboost_full_factor_bundle,
    )

    n = 120
    competitor = [80 + (i % 30) for i in range(n)]
    promo = [1.0 if i % 10 == 0 else 0.0 for i in range(n)]
    sales = [max(5.0, 220.0 - 1.2 * competitor[i] + 25.0 * promo[i]) for i in range(n)]
    daily = pd.DataFrame(
        {
            "date": pd.date_range("2024-01-01", periods=n),
            "sales": sales,
            "price": [100.0] * n,
            "discount": [0.0] * n,
            "net_unit_price": [100.0] * n,
            "freight_value": [5.0] * n,
            "cost": [60.0] * n,
            "promotion": promo,
            "review_score": [4.5] * n,
            "reviews_count": [10.0] * n,
            "stock": [999.0] * n,
            "factor__competitor_price": competitor,
        }
    )
    future = pd.DataFrame({"date": pd.date_range("2024-05-01", periods=14)})
    trained = train_catboost_full_factor_bundle(daily, future, min_train_days=60)
    assert "feature_cols" in trained
    bundle = {
        "daily_base": daily,
        "future_dates": future,
        "base_ctx": {"price": 100.0, "discount": 0.0, "promotion": 0.0, "freight_value": 5.0, "cost": 60.0},
        "catboost_full_factor_bundle": trained,
    }
    high_comp = predict_catboost_full_factor_projection(
        bundle,
        manual_price=100.0,
        horizon_days=14,
        overrides={"factor__competitor_price": 130.0},
    )
    low_comp = predict_catboost_full_factor_projection(
        bundle,
        manual_price=100.0,
        horizon_days=14,
        overrides={"factor__competitor_price": 80.0},
    )
    assert float(low_comp["demand_total"]) > float(high_comp["demand_total"])


def test_excel_contains_catboost_sheets():
    import app
    from openpyxl import load_workbook

    txn = _build_synthetic_txn()
    result = app.run_full_pricing_analysis_universal(
        normalized_txn=txn,
        target_category="cat1",
        target_sku="A",
        scenario_calc_mode="catboost_full_factors",
    )
    wb = load_workbook(result["excel_buffer"])
    assert "D_catboost_feature_report" in wb.sheetnames
    assert "D_catboost_importances" in wb.sheetnames
    assert "D_catboost_factor_catalog" in wb.sheetnames


def test_catboost_what_if_contract_fields():
    from catboost_full_factor_engine import train_catboost_full_factor_bundle, predict_catboost_full_factor_projection

    daily = pd.DataFrame(
        {
            "date": pd.date_range("2024-01-01", periods=90),
            "sales": [25 + (i % 6) for i in range(90)],
            "price": [100.0] * 90,
            "discount": [0.0] * 90,
            "net_unit_price": [100.0] * 90,
            "freight_value": [5.0] * 90,
            "cost": [60.0] * 90,
            "promotion": [0.0] * 90,
            "review_score": [4.5] * 90,
            "reviews_count": [10.0] * 90,
            "stock": [999.0] * 90,
            "factor__weather_index": [i % 4 for i in range(90)],
        }
    )
    future = pd.DataFrame({"date": pd.date_range("2024-04-01", periods=14)})
    cb = train_catboost_full_factor_bundle(daily, future, min_train_days=60)
    result = predict_catboost_full_factor_projection(
        trained_bundle={"daily_base": daily, "future_dates": future, "base_ctx": {"price": 100.0}, "catboost_full_factor_bundle": cb},
        manual_price=100.0,
        horizon_days=14,
    )
    assert result["scenario_calculation_path"] == "catboost_full_factor_reprediction"
    assert result["legacy_simulation_used"] is False
    assert result["model_backend"] == "catboost"


def test_catboost_holdout_metrics_are_recursive_mode():
    from catboost_full_factor_engine import train_catboost_full_factor_bundle

    daily = pd.DataFrame(
        {
            "date": pd.date_range("2024-01-01", periods=100),
            "sales": [15 + (i % 5) for i in range(100)],
            "price": [100.0 + (i % 4) for i in range(100)],
            "discount": [0.0] * 100,
            "freight_value": [5.0] * 100,
            "cost": [60.0] * 100,
            "promotion": [1.0 if i % 9 == 0 else 0.0 for i in range(100)],
            "stock": [999.0] * 100,
            "factor__competitor_price": [95.0 + (i % 8) for i in range(100)],
        }
    )
    future = pd.DataFrame({"date": pd.date_range("2024-04-15", periods=14)})
    bundle = train_catboost_full_factor_bundle(daily, future, min_train_days=60)
    assert bundle["holdout_metrics"]["mode"] == "recursive_daily_holdout"
    hp = bundle["holdout_predictions"]
    assert {"date", "actual_sales", "predicted_sales", "abs_error", "ape"}.issubset(set(hp.columns))


def test_safe_clip_and_extrapolation_have_different_financial_price_only():
    from catboost_full_factor_engine import normalize_price_guardrail_mode, train_catboost_full_factor_bundle, predict_catboost_full_factor_projection

    n = 100
    daily = pd.DataFrame(
        {
            "date": pd.date_range("2024-01-01", periods=n),
            "sales": [20 + (i % 5) for i in range(n)],
            "price": [18.0 + (i % 2) for i in range(n)],
            "discount": [0.0] * n,
            "net_unit_price": [18.0 + (i % 2) for i in range(n)],
            "freight_value": [4.0] * n,
            "cost": [11.0] * n,
            "promotion": [0.0] * n,
            "review_score": [4.6] * n,
            "reviews_count": [10.0] * n,
            "stock": [999.0] * n,
        }
    )
    future = pd.DataFrame({"date": pd.date_range("2024-05-01", periods=10)})
    cb = train_catboost_full_factor_bundle(daily, future, min_train_days=60)
    bundle = {"daily_base": daily, "future_dates": future, "base_ctx": {"price": 19.0, "discount": 0.0}, "catboost_full_factor_bundle": cb}
    wr_safe = predict_catboost_full_factor_projection(bundle, manual_price=35.0, horizon_days=5, price_guardrail_mode="safe_clip")
    wr_extra = predict_catboost_full_factor_projection(bundle, manual_price=35.0, horizon_days=5, price_guardrail_mode="economic_extrapolation")
    assert wr_safe["price_guardrail_mode"] == "safe_clip"
    assert wr_extra["price_guardrail_mode"] == "economic_extrapolation"
    assert wr_safe["model_price"] == pytest.approx(wr_safe["safe_price_gross"])
    assert wr_extra["model_price"] == pytest.approx(wr_extra["safe_price_gross"])
    assert wr_safe["applied_price_gross"] == pytest.approx(wr_safe["safe_price_gross"])
    assert wr_extra["applied_price_gross"] == pytest.approx(35.0)
    assert wr_safe["extrapolation_applied"] is False
    assert wr_extra["extrapolation_applied"] is True
    assert normalize_price_guardrail_mode("exact_manual") == "economic_extrapolation"
    assert normalize_price_guardrail_mode("strict") == "economic_extrapolation"
    assert normalize_price_guardrail_mode("manual") == "economic_extrapolation"


def test_safe_clip_behavior_is_unchanged_for_out_of_range_price():
    from catboost_full_factor_engine import train_catboost_full_factor_bundle, predict_catboost_full_factor_projection
    n = 100
    daily = pd.DataFrame({"date": pd.date_range("2024-01-01", periods=n), "sales": [20 + (i % 5) for i in range(n)], "price": [18.0 + (i % 2) for i in range(n)], "discount": [0.0] * n, "net_unit_price": [18.0 + (i % 2) for i in range(n)], "freight_value": [4.0] * n, "cost": [11.0] * n, "promotion": [0.0] * n, "review_score": [4.6] * n, "reviews_count": [10.0] * n, "stock": [999.0] * n})
    future = pd.DataFrame({"date": pd.date_range("2024-05-01", periods=10)})
    bundle = {"daily_base": daily, "future_dates": future, "base_ctx": {"price": 19.0, "discount": 0.0}, "catboost_full_factor_bundle": train_catboost_full_factor_bundle(daily, future, min_train_days=60)}
    wr = predict_catboost_full_factor_projection(bundle, manual_price=40.99, horizon_days=5, price_guardrail_mode="safe_clip")
    assert wr["price_guardrail_mode"] == "safe_clip"
    assert wr["requested_price"] == pytest.approx(40.99)
    assert wr["price_out_of_range"] is True
    assert wr["price_clipped"] is True
    assert wr["clip_applied"] is True
    assert wr["extrapolation_applied"] is False
    assert wr["model_price"] == pytest.approx(wr["safe_price_gross"])
    assert wr["price_for_model"] == pytest.approx(wr["safe_price_gross"])
    assert wr["applied_price_gross"] == pytest.approx(wr["safe_price_gross"])
    assert wr["scenario_price_effect_source"] == "catboost_full_factor_reprediction"
    assert not np.isfinite(wr.get("elasticity_used", np.nan))
    assert wr.get("extrapolation_tail_multiplier", 1.0) == pytest.approx(1.0)


def test_economic_extrapolation_inside_range_does_not_apply_extrapolation():
    from catboost_full_factor_engine import train_catboost_full_factor_bundle, predict_catboost_full_factor_projection
    n = 90
    daily = pd.DataFrame({"date": pd.date_range("2024-01-01", periods=n), "sales": [10 + i % 3 for i in range(n)], "price": [20.0] * n, "discount": [0.0] * n, "net_unit_price": [20.0] * n, "freight_value": [2.0] * n, "cost": [12.0] * n, "promotion": [0.0] * n, "review_score": [4.5] * n, "reviews_count": [1.0] * n, "stock": [999.0] * n})
    future = pd.DataFrame({"date": pd.date_range("2024-05-01", periods=3)})
    wr = predict_catboost_full_factor_projection({"daily_base": daily, "future_dates": future, "base_ctx": {"price": 20.0}, "catboost_full_factor_bundle": train_catboost_full_factor_bundle(daily, future, min_train_days=60)}, manual_price=19.99, horizon_days=2, price_guardrail_mode="economic_extrapolation")
    assert wr["price_guardrail_mode"] == "economic_extrapolation"
    assert wr["price_out_of_range"] is False
    assert wr["price_clipped"] is False
    assert wr["extrapolation_applied"] is False
    assert wr["model_price"] == pytest.approx(19.99)
    assert wr["price_for_model"] == pytest.approx(19.99)
    assert wr["applied_price_gross"] == pytest.approx(19.99)


def test_catboost_daily_clip_applied_is_exported_for_safe_clip():
    from catboost_full_factor_engine import train_catboost_full_factor_bundle, predict_catboost_full_factor_projection
    n = 90
    daily = pd.DataFrame({"date": pd.date_range("2024-01-01", periods=n), "sales": [10 + i % 3 for i in range(n)], "price": [20.0] * n, "discount": [0.0] * n, "net_unit_price": [20.0] * n, "freight_value": [2.0] * n, "cost": [12.0] * n, "promotion": [0.0] * n, "review_score": [4.5] * n, "reviews_count": [1.0] * n, "stock": [999.0] * n})
    future = pd.DataFrame({"date": pd.date_range("2024-05-01", periods=3)})
    wr = predict_catboost_full_factor_projection({"daily_base": daily, "future_dates": future, "base_ctx": {"price": 20.0}, "catboost_full_factor_bundle": train_catboost_full_factor_bundle(daily, future, min_train_days=60)}, manual_price=35.0, horizon_days=2, price_guardrail_mode="safe_clip")
    assert wr["daily"]["clip_applied"].iloc[0] in [True, 1]


def test_applied_path_summary_uses_financial_price_in_extrapolation():
    from catboost_full_factor_engine import train_catboost_full_factor_bundle, predict_catboost_full_factor_projection
    n = 120
    daily = pd.DataFrame({"date": pd.date_range("2024-01-01", periods=n), "sales": [10 + i % 3 for i in range(n)], "price": [20.0] * n, "discount": [0.1] * n, "net_unit_price": [18.0] * n, "freight_value": [2.0] * n, "cost": [12.0] * n, "promotion": [0.0] * n, "review_score": [4.5] * n, "reviews_count": [1.0] * n, "stock": [999.0] * n})
    future = pd.DataFrame({"date": pd.date_range("2024-05-01", periods=5)})
    wr = predict_catboost_full_factor_projection({"daily_base": daily, "future_dates": future, "base_ctx": {"price": 20.0, "discount": 0.1}, "catboost_full_factor_bundle": train_catboost_full_factor_bundle(daily, future, min_train_days=60)}, manual_price=35.0, horizon_days=3, price_guardrail_mode="economic_extrapolation")
    assert wr["applied_path_summary"]["price_net_avg"] == pytest.approx(wr["applied_price_net"])
    assert wr["applied_path_summary"]["price_net_avg"] != pytest.approx(wr["model_price"] * (1.0 - 0.1))


def test_low_price_extrapolation_increases_or_keeps_tail_multiplier_above_one():
    from catboost_full_factor_engine import train_catboost_full_factor_bundle, predict_catboost_full_factor_projection
    n = 120
    daily = pd.DataFrame({"date": pd.date_range("2024-01-01", periods=n), "sales": [10 + i % 3 for i in range(n)], "price": [20.0] * n, "discount": [0.0] * n, "net_unit_price": [20.0] * n, "freight_value": [2.0] * n, "cost": [12.0] * n, "promotion": [0.0] * n, "review_score": [4.5] * n, "reviews_count": [1.0] * n, "stock": [999.0] * n})
    future = pd.DataFrame({"date": pd.date_range("2024-05-01", periods=5)})
    very_low_price = 5.0
    wr = predict_catboost_full_factor_projection({"daily_base": daily, "future_dates": future, "base_ctx": {"price": 20.0}, "catboost_full_factor_bundle": train_catboost_full_factor_bundle(daily, future, min_train_days=60)}, manual_price=very_low_price, horizon_days=3, price_guardrail_mode="economic_extrapolation")
    assert wr["extrapolation_applied"] is True
    assert wr["applied_price_gross"] == pytest.approx(very_low_price)
    assert wr["model_price"] == pytest.approx(wr["safe_price_gross"])
    assert wr["extrapolation_tail_multiplier"] >= 1.0


def test_high_price_extrapolation_reduces_or_keeps_tail_multiplier_below_one():
    from catboost_full_factor_engine import train_catboost_full_factor_bundle, predict_catboost_full_factor_projection
    n = 120
    daily = pd.DataFrame({"date": pd.date_range("2024-01-01", periods=n), "sales": [10 + i % 3 for i in range(n)], "price": [20.0] * n, "discount": [0.0] * n, "net_unit_price": [20.0] * n, "freight_value": [2.0] * n, "cost": [12.0] * n, "promotion": [0.0] * n, "review_score": [4.5] * n, "reviews_count": [1.0] * n, "stock": [999.0] * n})
    future = pd.DataFrame({"date": pd.date_range("2024-05-01", periods=5)})
    wr = predict_catboost_full_factor_projection({"daily_base": daily, "future_dates": future, "base_ctx": {"price": 20.0}, "catboost_full_factor_bundle": train_catboost_full_factor_bundle(daily, future, min_train_days=60)}, manual_price=35.0, horizon_days=3, price_guardrail_mode="economic_extrapolation")
    assert wr["extrapolation_tail_multiplier"] <= 1.0


def test_default_price_guardrail_mode_is_safe_clip():
    from catboost_full_factor_engine import train_catboost_full_factor_bundle, predict_catboost_full_factor_projection

    n = 90
    daily = pd.DataFrame({"date": pd.date_range("2024-01-01", periods=n), "sales": [10 + i % 3 for i in range(n)], "price": [20.0] * n, "discount": [0.0] * n, "net_unit_price": [20.0] * n, "freight_value": [2.0] * n, "cost": [12.0] * n, "promotion": [0.0] * n, "review_score": [4.5] * n, "reviews_count": [1.0] * n, "stock": [999.0] * n})
    future = pd.DataFrame({"date": pd.date_range("2024-05-01", periods=3)})
    cb = train_catboost_full_factor_bundle(daily, future, min_train_days=60)
    wr = predict_catboost_full_factor_projection({"daily_base": daily, "future_dates": future, "base_ctx": {"price": 20.0}, "catboost_full_factor_bundle": cb}, manual_price=35.0, horizon_days=2)
    assert wr["price_guardrail_mode"] == "safe_clip"


def test_catboost_daily_preserves_shock_multiplier():
    from catboost_full_factor_engine import train_catboost_full_factor_bundle, predict_catboost_full_factor_projection

    n = 90
    daily = pd.DataFrame({"date": pd.date_range("2024-01-01", periods=n), "sales": [10 + i % 3 for i in range(n)], "price": [20.0] * n, "discount": [0.0] * n, "net_unit_price": [20.0] * n, "freight_value": [2.0] * n, "cost": [12.0] * n, "promotion": [0.0] * n, "review_score": [4.5] * n, "reviews_count": [1.0] * n, "stock": [999.0] * n})
    future = pd.DataFrame({"date": pd.date_range("2024-05-01", periods=3)})
    cb = train_catboost_full_factor_bundle(daily, future, min_train_days=60)
    wr = predict_catboost_full_factor_projection({"daily_base": daily, "future_dates": future, "base_ctx": {"price": 20.0}, "catboost_full_factor_bundle": cb}, manual_price=20.0, horizon_days=2, demand_multiplier=0.87)
    assert wr["daily"]["shock_multiplier"].iloc[0] == pytest.approx(0.87)


def test_extra_factor_no_backward_fill_in_daily_builder():
    from data_adapter import build_daily_from_transactions

    src = pd.DataFrame(
        [
            {"date": "2025-01-01", "product_id": "sku-1", "category": "cat-a", "price": 100, "quantity": 1, "revenue": 100, "custom_factor": None},
            {"date": "2025-01-03", "product_id": "sku-1", "category": "cat-a", "price": 100, "quantity": 1, "revenue": 100, "custom_factor": 10.0},
            {"date": "2025-01-04", "product_id": "sku-1", "category": "cat-a", "price": 100, "quantity": 1, "revenue": 100, "custom_factor": 12.0},
        ]
    )
    src["date"] = pd.to_datetime(src["date"])

    daily = build_daily_from_transactions(src, "sku-1", target_category="cat-a", include_extra_factors=True)

    first = daily[daily["date"] == pd.Timestamp("2025-01-01")].iloc[0]
    mid = daily[daily["date"] == pd.Timestamp("2025-01-02")].iloc[0]
    third = daily[daily["date"] == pd.Timestamp("2025-01-03")].iloc[0]
    fourth = daily[daily["date"] == pd.Timestamp("2025-01-04")].iloc[0]

    assert float(first["factor__custom_factor"]) == 0.0
    assert float(mid["factor__custom_factor"]) == 0.0
    assert float(third["factor__custom_factor"]) == 10.0
    assert float(fourth["factor__custom_factor"]) == 12.0
    assert int(first["factor__custom_factor__was_missing"]) == 1
    assert int(mid["factor__custom_factor__was_missing"]) == 1
    assert int(third["factor__custom_factor__was_missing"]) == 0
    assert int(fourth["factor__custom_factor__was_missing"]) == 0


def test_catboost_train_without_is_stockout_column_does_not_crash():
    pytest.importorskip("catboost")
    from catboost_full_factor_engine import train_catboost_full_factor_bundle

    n = 90
    daily = pd.DataFrame({
        "date": pd.date_range("2024-01-01", periods=n),
        "sales": [20 + (i % 5) for i in range(n)],
        "price": [100.0 + (i % 4) for i in range(n)],
        "discount": [0.0] * n,
        "net_unit_price": [100.0 + (i % 4) for i in range(n)],
        "freight_value": [5.0] * n,
        "cost": [65.0] * n,
        "promotion": [0.0, 1.0] * (n // 2),
        "review_score": [4.5] * n,
        "reviews_count": [10.0] * n,
        "stock": [999.0] * n,
    })
    future = pd.DataFrame({"date": pd.date_range(daily["date"].max() + pd.Timedelta(days=1), periods=14)})
    bundle = train_catboost_full_factor_bundle(daily, future, min_train_days=60)
    assert "target_semantics" in bundle
    assert float(bundle.get("target_semantics", {}).get("stockout_share", 0.0)) == 0.0
